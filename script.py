import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Thresholds for investment decision
THRESHOLDS = {
    "P/E Ratio": 20,
    "P/B Ratio": 3,
    "Dividend Yield (%)": 2,
    "Debt to Equity": 1,
    "Return on Equity (%)": 15,
    "Earnings Yield (%)": 5
}

# Fetch financial data for a single ticker
def fetch_financial_data(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        # Calculate metrics
        pe_ratio = info.get("trailingPE", "N/A")
        pb_ratio = info.get("priceToBook", "N/A")
        dividend_yield = info.get("dividendYield", 0) * 100 if info.get("dividendYield") else "N/A"
        market_cap = info.get("marketCap", "N/A")
        roe = info.get("returnOnEquity", 0) * 100 if info.get("returnOnEquity") else "N/A"
        earnings_yield = (100 / pe_ratio) if pe_ratio and pe_ratio != "N/A" else "N/A"
        debt_to_equity = info.get("debtToEquity", "N/A")

        # Investment decision
        invest = "Yes" if (
            (pe_ratio != "N/A" and pe_ratio <= THRESHOLDS["P/E Ratio"])
            and (pb_ratio != "N/A" and pb_ratio <= THRESHOLDS["P/B Ratio"])
            and (dividend_yield != "N/A" and dividend_yield >= THRESHOLDS["Dividend Yield (%)"])
            and (roe != "N/A" and roe >= THRESHOLDS["Return on Equity (%)"])
            and (earnings_yield != "N/A" and earnings_yield >= THRESHOLDS["Earnings Yield (%)"])
        ) else "No"

        return {
            "Ticker": ticker,
            "P/E Ratio": pe_ratio,
            "P/B Ratio": pb_ratio,
            "Dividend Yield (%)": dividend_yield,
            "Market Cap": market_cap,
            "Return on Equity (%)": roe,
            "Earnings Yield (%)": earnings_yield,
            "Debt to Equity": debt_to_equity,
            "Invest": invest
        }
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
        return None

# Save results to a styled Excel file
def save_to_styled_excel(data, output_file):
    # Convert data to DataFrame
    df = pd.DataFrame(data)

    # Create a workbook and add data
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"

    # Add header row
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = alignment

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

# Main function
def main():
    tickers = [
    "TSLA", "AAPL", "MSFT", "AMZN", "GOOGL", "META", "NVDA", "BRK-B", "V", "JPM",
    "JNJ", "WMT", "PG", "XOM", "UNH", "HD", "MA", "PFE", "CVX", "KO",
    "DIS", "PEP", "BAC", "CSCO", "NFLX", "ADBE", "CMCSA", "ABT", "INTC", "PYPL",
    "T", "MRK", "ORCL", "NKE", "LLY", "IBM", "CRM", "ACN", "MDT", "DHR",
    "BMY", "VZ", "MCD", "ABBV", "NEE", "COST", "TXN", "PM", "UNP", "UPS",
    "LOW", "HON", "RTX", "SBUX", "SPGI", "MS", "INTU", "QCOM", "CAT", "AXP",
    "AMGN", "LIN", "ZTS", "BLK", "TMO", "GILD", "CVS", "BA", "DE", "PLD",
    "NOW", "MMM", "ADP", "C", "BKNG", "ISRG", "REGN", "EL", "VRTX", "MO",
    "LRCX", "F", "GM", "TWTR", "SNPS", "ADI", "SYK", "EW", "ETN", "PGR",
    "WBA", "CL", "BSX", "DG", "HUM", "TGT", "APD", "CI", "GD", "FDX"
    ]
    output_file = "Stock_Analysis.xlsx"

    # Fetch data for each ticker
    results = []
    for ticker in tickers:
        print(f"Processing {ticker}...")
        data = fetch_financial_data(ticker)
        if data:
            results.append(data)

    # Save to styled Excel
    save_to_styled_excel(results, output_file)

if __name__ == "__main__":
    main()
